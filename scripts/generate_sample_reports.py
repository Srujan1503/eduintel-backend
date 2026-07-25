"""
Integration test to generate sample exports and verify they can be opened/parsed.
Run this to verify all export formats work correctly.
"""

import uuid
from datetime import date, datetime
from decimal import Decimal
from types import SimpleNamespace

from app.services.report_service import ReportService

# Create mock database and objects
class MockDB:
    def __init__(self):
        self.school = None
        self.campaigns = []
        self.competitors = []

    def query(self, model):
        return MockQuery(model, self)


class MockQuery:
    def __init__(self, model, db):
        self.model = model
        self.db = db
        self.filters_applied = []

    def filter(self, *conditions):
        self.filters_applied.extend(conditions)
        return self

    def first(self):
        if self.model.__name__ == "School":
            return self.db.school
        return None

    def all(self):
        if self.model.__name__ == "Campaign":
            return self.db.campaigns
        elif self.model.__name__ == "Competitor":
            return self.db.competitors
        return []


db = MockDB()

# Create school
db.school = SimpleNamespace(
    id=uuid.uuid4(),
    name="Demo School",
    type="school",
    subscription_tier="premium",
    city="San Francisco",
    state="CA",
    country="USA",
    is_active=True,
)

# Create campaigns
for i in range(3):
    campaign = SimpleNamespace(
        id=uuid.uuid4(),
        school_id=db.school.id,
        name=f"Q{i+1} Marketing Campaign",
        channel=["email", "social", "search"][i % 3],
        start_date=date(2026, i * 3 + 1, 1),
        end_date=date(2026, i * 3 + 3, 30),
        budget=Decimal(str(10000 + i * 5000)),
        spend=Decimal(str(7000 + i * 3000)),
        conversions=150 + i * 50,
        meta=None,
    )
    db.campaigns.append(campaign)

# Create competitors
for i in range(5):
    competitor = SimpleNamespace(
        id=uuid.uuid4(),
        school_id=db.school.id,
        name=f"Competitor {i+1}",
        domain=f"competitor{i+1}.com",
        threat_score=float(0.3 + i * 0.12),
        first_seen=datetime(2026, 1, 1),
        last_seen=datetime(2026, 7, 25),
        meta=None,
    )
    db.competitors.append(competitor)

# Generate reports
service = ReportService(db)

# CSV
csv_content = service.generate_csv(db.school.id)
with open("sample_report.csv", "w") as f:
    f.write(csv_content)
print(f"✓ CSV export created: sample_report.csv ({len(csv_content)} bytes)")

# Excel
excel_bytes = service.generate_excel(db.school.id)
with open("sample_report.xlsx", "wb") as f:
    f.write(excel_bytes)
print(f"✓ Excel export created: sample_report.xlsx ({len(excel_bytes)} bytes)")

# PDF
pdf_bytes = service.generate_pdf(db.school.id)
with open("sample_report.pdf", "wb") as f:
    f.write(pdf_bytes)
print(f"✓ PDF export created: sample_report.pdf ({len(pdf_bytes)} bytes)")

# Verify file formats
import csv
from io import StringIO

# Verify CSV
reader = csv.reader(StringIO(csv_content))
rows = list(reader)
assert len(rows) > 10, "CSV should have multiple rows"
assert "Demo School" in csv_content, "CSV should contain school name"
print(f"✓ CSV verified: {len(rows)} rows")

# Verify Excel
import openpyxl

wb = openpyxl.load_workbook("sample_report.xlsx")
ws = wb.active
assert ws["A1"].value == "Report for Demo School"
print(f"✓ Excel verified: {ws.max_row} rows")

# Verify PDF
assert pdf_bytes[:4] == b"%PDF", "PDF should start with PDF signature"
print(f"✓ PDF verified: valid PDF format ({len(pdf_bytes)} bytes)")

print("\n✓ All export formats verified successfully!")

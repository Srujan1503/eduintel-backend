import csv
import io
from datetime import datetime, date
from typing import Optional

from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.school import School
from app.models.campaign import Campaign
from app.models.competitor import Competitor


class ReportData:
    """Container for report data with metadata."""

    def __init__(self, school_id, school_name: str, generated_at: datetime):
        self.school_id = school_id
        self.school_name = school_name
        self.generated_at = generated_at
        self.campaigns = []
        self.competitors = []
        self.summary = {}


class ReportService:
    """Service for generating marketing intelligence reports in multiple formats."""

    def __init__(self, db: Session):
        self.db = db

    def gather_report_data(
        self,
        school_id,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> ReportData:
        """Fetch all report data for a school with optional date filtering."""
        school = self.db.query(School).filter(School.id == school_id).first()
        if not school:
            return None

        report_data = ReportData(
            school_id=school_id, school_name=school.name, generated_at=datetime.now()
        )

        # Fetch campaigns with optional date filtering
        campaigns_query = self.db.query(Campaign).filter(Campaign.school_id == school_id, Campaign.deleted_at.is_(None))
        if start_date:
            campaigns_query = campaigns_query.filter(Campaign.start_date >= start_date)
        if end_date:
            campaigns_query = campaigns_query.filter(Campaign.end_date <= end_date)
        report_data.campaigns = campaigns_query.all()

        # Fetch competitors
        report_data.competitors = (
            self.db.query(Competitor).filter(Competitor.school_id == school_id, Competitor.deleted_at.is_(None)).all()
        )

        # Calculate summary statistics
        report_data.summary = {
            "total_campaigns": len(report_data.campaigns),
            "total_competitors": len(report_data.competitors),
            "total_budget": sum(float(c.budget or 0) for c in report_data.campaigns),
            "total_spend": sum(float(c.spend or 0) for c in report_data.campaigns),
            "total_conversions": sum(int(c.conversions or 0) for c in report_data.campaigns),
            "avg_threat_score": (
                sum(c.threat_score or 0.0 for c in report_data.competitors) / len(report_data.competitors)
                if report_data.competitors
                else 0.0
            ),
        }

        return report_data

    def generate_csv(
        self,
        school_id,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> str:
        """Generate a CSV report with campaign and competitor data."""
        report_data = self.gather_report_data(school_id, start_date, end_date)
        if not report_data:
            return None

        output = io.StringIO()
        writer = csv.writer(output)

        # Write header with school name
        writer.writerow([f"Report for {report_data.school_name}"])
        writer.writerow([f"Generated: {report_data.generated_at.isoformat()}"])
        writer.writerow([])

        # Write summary statistics
        writer.writerow(["Summary Statistics"])
        writer.writerow(["Total Campaigns", report_data.summary["total_campaigns"]])
        writer.writerow(["Total Competitors", report_data.summary["total_competitors"]])
        writer.writerow(["Total Budget", f"${report_data.summary['total_budget']:.2f}"])
        writer.writerow(["Total Spend", f"${report_data.summary['total_spend']:.2f}"])
        writer.writerow(["Total Conversions", report_data.summary["total_conversions"]])
        writer.writerow(["Average Threat Score", f"{report_data.summary['avg_threat_score']:.2f}"])
        writer.writerow([])

        # Write campaigns section
        writer.writerow(["Campaigns"])
        writer.writerow(
            [
                "ID",
                "Name",
                "Channel",
                "Start Date",
                "End Date",
                "Budget",
                "Spend",
                "Conversions",
            ]
        )
        for campaign in report_data.campaigns:
            writer.writerow(
                [
                    str(campaign.id),
                    campaign.name,
                    campaign.channel or "",
                    campaign.start_date or "",
                    campaign.end_date or "",
                    f"${float(campaign.budget or 0):.2f}",
                    f"${float(campaign.spend or 0):.2f}",
                    campaign.conversions or 0,
                ]
            )
        writer.writerow([])

        # Write competitors section
        writer.writerow(["Competitors"])
        writer.writerow(["ID", "Name", "Domain", "Threat Score", "First Seen", "Last Seen"])
        for competitor in report_data.competitors:
            writer.writerow(
                [
                    str(competitor.id),
                    competitor.name,
                    competitor.domain or "",
                    competitor.threat_score or 0.0,
                    competitor.first_seen or "",
                    competitor.last_seen or "",
                ]
            )

        return output.getvalue()

    def generate_excel(
        self,
        school_id,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Generate an Excel workbook with campaign and competitor data."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        report_data = self.gather_report_data(school_id, start_date, end_date)
        if not report_data:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        row = 1

        # Title
        ws[f"A{row}"] = f"Report for {report_data.school_name}"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        ws[f"A{row}"] = f"Generated: {report_data.generated_at.isoformat()}"
        row += 2

        # Summary section
        ws[f"A{row}"] = "Summary Statistics"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        summary_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        headers_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        ws[f"A{row}"] = "Metric"
        ws[f"B{row}"] = "Value"
        ws[f"A{row}"].fill = headers_fill
        ws[f"B{row}"].fill = headers_fill
        row += 1

        summary_items = [
            ("Total Campaigns", report_data.summary["total_campaigns"]),
            ("Total Competitors", report_data.summary["total_competitors"]),
            ("Total Budget", f"${report_data.summary['total_budget']:.2f}"),
            ("Total Spend", f"${report_data.summary['total_spend']:.2f}"),
            ("Total Conversions", report_data.summary["total_conversions"]),
            ("Average Threat Score", f"{report_data.summary['avg_threat_score']:.2f}"),
        ]

        for label, value in summary_items:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].fill = summary_fill
            ws[f"B{row}"].fill = summary_fill
            row += 1

        row += 1

        # Campaigns section
        ws[f"A{row}"] = "Campaigns"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        campaign_headers = ["ID", "Name", "Channel", "Start Date", "End Date", "Budget", "Spend", "Conversions"]
        for col, header in enumerate(campaign_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = headers_fill
            cell.font = Font(bold=True)

        row += 1

        for campaign in report_data.campaigns:
            ws.cell(row=row, column=1).value = str(campaign.id)
            ws.cell(row=row, column=2).value = campaign.name
            ws.cell(row=row, column=3).value = campaign.channel or ""
            ws.cell(row=row, column=4).value = campaign.start_date
            ws.cell(row=row, column=5).value = campaign.end_date
            ws.cell(row=row, column=6).value = float(campaign.budget or 0)
            ws.cell(row=row, column=6).number_format = "$#,##0.00"
            ws.cell(row=row, column=7).value = float(campaign.spend or 0)
            ws.cell(row=row, column=7).number_format = "$#,##0.00"
            ws.cell(row=row, column=8).value = campaign.conversions or 0
            row += 1

        # Auto-size columns
        for col in range(1, len(campaign_headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        row += 1

        # Competitors section
        ws[f"A{row}"] = "Competitors"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        competitor_headers = ["ID", "Name", "Domain", "Threat Score", "First Seen", "Last Seen"]
        for col, header in enumerate(competitor_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = headers_fill
            cell.font = Font(bold=True)

        row += 1

        for competitor in report_data.competitors:
            ws.cell(row=row, column=1).value = str(competitor.id)
            ws.cell(row=row, column=2).value = competitor.name
            ws.cell(row=row, column=3).value = competitor.domain or ""
            ws.cell(row=row, column=4).value = competitor.threat_score or 0.0
            ws.cell(row=row, column=5).value = competitor.first_seen
            ws.cell(row=row, column=6).value = competitor.last_seen
            row += 1

        # Auto-size competitor columns
        for col in range(1, len(competitor_headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        # Convert workbook to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_pdf(
        self,
        school_id,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Generate a professional PDF report with campaign and competitor data."""
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

        report_data = self.gather_report_data(school_id, start_date, end_date)
        if not report_data:
            return None

        # Create PDF document
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=0.5 * inch, leftMargin=0.5 * inch)
        story = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Heading1"], fontSize=24, textColor=colors.HexColor("#1F4E78"), spaceAfter=12
        )
        subtitle_style = ParagraphStyle(
            "CustomSubtitle", parent=styles["Heading2"], fontSize=12, textColor=colors.grey, spaceAfter=6
        )

        # Title
        story.append(Paragraph(f"Marketing Intelligence Report", title_style))
        story.append(Paragraph(f"School: {report_data.school_name}", subtitle_style))
        story.append(Paragraph(f"Generated: {report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
        story.append(Spacer(1, 0.3 * inch))

        # Summary section
        story.append(Paragraph("Summary Statistics", styles["Heading2"]))
        summary_data = [
            ["Metric", "Value"],
            ["Total Campaigns", str(report_data.summary["total_campaigns"])],
            ["Total Competitors", str(report_data.summary["total_competitors"])],
            ["Total Budget", f"${report_data.summary['total_budget']:.2f}"],
            ["Total Spend", f"${report_data.summary['total_spend']:.2f}"],
            ["Total Conversions", str(report_data.summary["total_conversions"])],
            ["Average Threat Score", f"{report_data.summary['avg_threat_score']:.2f}"],
        ]

        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 11),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(summary_table)
        story.append(Spacer(1, 0.3 * inch))

        # Campaigns section
        if report_data.campaigns:
            story.append(Paragraph("Campaigns", styles["Heading2"]))
            campaign_data = [["ID", "Name", "Channel", "Budget", "Spend", "Conversions"]]
            for campaign in report_data.campaigns:
                campaign_data.append(
                    [
                        str(campaign.id)[:8],
                        campaign.name[:20],
                        campaign.channel or "N/A",
                        f"${float(campaign.budget or 0):.0f}",
                        f"${float(campaign.spend or 0):.0f}",
                        str(campaign.conversions or 0),
                    ]
                )

            campaign_table = Table(campaign_data, colWidths=[1 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1 * inch, 1 * inch])
            campaign_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )
            story.append(campaign_table)
            story.append(Spacer(1, 0.3 * inch))

        # Competitors section
        if report_data.competitors:
            story.append(Paragraph("Competitors", styles["Heading2"]))
            competitor_data = [["Name", "Domain", "Threat Score"]]
            for competitor in report_data.competitors:
                competitor_data.append(
                    [
                        competitor.name[:25],
                        competitor.domain or "N/A",
                        f"{competitor.threat_score or 0.0:.2f}",
                    ]
                )

            competitor_table = Table(competitor_data, colWidths=[2.5 * inch, 2 * inch, 1.5 * inch])
            competitor_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )
            story.append(competitor_table)

        # Build PDF
        doc.build(story)
        output.seek(0)
        return output.getvalue()
